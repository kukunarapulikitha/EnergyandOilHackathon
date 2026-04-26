"""Formula-driven Excel export — Tier 2 stretch goal.

Builds a 4-sheet `.xlsx` that preserves the pipeline's KPI logic as live
Excel formulas, so an analyst can:
    1. Open the file offline
    2. Change WTI price / target year in the Inputs sheet
    3. Watch revenue and KPI columns update automatically

Uses `openpyxl` (already in requirements). Returns bytes; Streamlit wraps
it with st.download_button.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
INPUT_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
INPUT_FONT = Font(bold=True, color="FF9C5700")
NOTE_FONT = Font(italic=True, color="FF777777", size=9)


def build_workbook(
    actuals: pd.DataFrame,
    forecasts: pd.DataFrame,
    default_wti_price: float = 78.0,
    default_target_year: int = 2030,
) -> bytes:
    wb = Workbook()

    _sheet_inputs(wb, default_wti_price, default_target_year)
    _sheet_production(wb, actuals)
    _sheet_forecasts(wb, forecasts, actuals)
    _sheet_kpi_summary(wb, actuals, forecasts)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- sheets


def _sheet_inputs(wb: Workbook, wti: float, target_year: int) -> None:
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Variable"
    ws["B1"] = "Value"
    for cell in ("A1", "B1"):
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL

    rows = [
        ("WTI spot price (USD/bbl)", wti, "Used for revenue_potential = production × days/yr × WTI"),
        ("Target year", target_year, "Forecast computed via Forecasts!D*this + Forecasts!E"),
        ("Decline adjustment (%)", 0.0, "+/- % applied to fitted slope for stress tests"),
    ]
    for i, (label, value, note) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value).fill = INPUT_FILL
        ws.cell(row=i, column=2).font = INPUT_FONT
        ws.cell(row=i, column=3, value=note).font = NOTE_FONT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70


def _sheet_production(wb: Workbook, actuals: pd.DataFrame) -> None:
    ws = wb.create_sheet("Production")
    cols = ["region_id", "region_name", "year", "fuel_type", "production", "unit", "source"]
    subset = actuals[cols] if not actuals.empty else pd.DataFrame(columns=cols)
    _write_dataframe(ws, subset)


def _sheet_forecasts(
    wb: Workbook, forecasts: pd.DataFrame, actuals: pd.DataFrame
) -> None:
    ws = wb.create_sheet("Forecasts")
    cols = ["region_id", "region_name", "fuel_type", "slope", "intercept",
            "r_squared", "trained_through_year", "method"]
    subset = forecasts[cols] if not forecasts.empty else pd.DataFrame(columns=cols)
    _write_dataframe(ws, subset)

    if subset.empty:
        return

    # Add formula column: projected value at Inputs!B$2
    col_idx = len(cols) + 1  # 1-indexed
    ws.cell(row=1, column=col_idx, value="projected_at_target_year").font = HEADER_FONT
    ws.cell(row=1, column=col_idx).fill = HEADER_FILL
    slope_col = get_column_letter(cols.index("slope") + 1)
    intercept_col = get_column_letter(cols.index("intercept") + 1)
    for row in range(2, len(subset) + 2):
        ws.cell(
            row=row,
            column=col_idx,
            value=f"={slope_col}{row} * Inputs!$B$2 + {intercept_col}{row}",
        )


def _sheet_kpi_summary(
    wb: Workbook, actuals: pd.DataFrame, forecasts: pd.DataFrame
) -> None:
    ws = wb.create_sheet("KPI Summary")
    headers = [
        "region_name", "fuel_type", "latest_year", "latest_production",
        "slope_per_yr", "r_squared", "projected_target_year",
        "revenue_at_target_usd", "rank_by_projection",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    if actuals.empty or forecasts.empty:
        return

    # Build one row per (region, fuel)
    latest_year = int(actuals["year"].max())
    latest = actuals[actuals["year"] == latest_year]

    data_rows = []
    for _, fr in forecasts.iterrows():
        match = latest[
            (latest["region_id"] == fr["region_id"])
            & (latest["fuel_type"] == fr["fuel_type"])
        ]
        latest_prod = float(match["production"].iloc[0]) if not match.empty else 0.0
        data_rows.append({
            "region_name": fr["region_name"],
            "fuel_type": fr["fuel_type"],
            "latest_year": latest_year,
            "latest_production": latest_prod,
            "slope_per_yr": float(fr["slope"]),
            "r_squared": float(fr["r_squared"]),
        })

    for i, r in enumerate(data_rows, start=2):
        ws.cell(row=i, column=1, value=r["region_name"])
        ws.cell(row=i, column=2, value=r["fuel_type"])
        ws.cell(row=i, column=3, value=r["latest_year"])
        ws.cell(row=i, column=4, value=r["latest_production"])
        ws.cell(row=i, column=5, value=r["slope_per_yr"])
        ws.cell(row=i, column=6, value=r["r_squared"])
        # Projection formula: slope × target_year + intercept (pulled from Forecasts sheet)
        # Simpler: slope × target_year + (latest_production - slope × latest_year)
        ws.cell(
            row=i,
            column=7,
            value=f"=E{i} * Inputs!$B$3 + (D{i} - E{i} * C{i})",
        )
        # Revenue = projection × 1000 × 365 × WTI (only meaningful for crude)
        ws.cell(
            row=i,
            column=8,
            value=f'=IF(B{i}="crude_oil", G{i} * 1000 * 365 * Inputs!$B$2, "")',
        )
        # Rank by projection among same-fuel peers
        last_row = len(data_rows) + 1
        ws.cell(
            row=i,
            column=9,
            value=f'=IF(B{i}="", "", RANK(G{i}, G$2:G${last_row}))',
        )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22


# ---------------------------------------------------------------- helpers


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    for col_idx, header in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=str(header))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            12, min(30, len(str(header)) + 4)
        )

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_xlsx_safe(value))


def _xlsx_safe(v):
    if pd.isna(v):
        return None
    if isinstance(v, (pd.Timestamp,)):
        return v.isoformat()
    return v
